from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Tuple

from fastapi import FastAPI, HTTPException
from pydantic import BaseModel, Field
from openpyxl import load_workbook

from .server import ExcelMCPServer


# -----------------------------
# Workbook Provider
# -----------------------------
class WorkbookProvider:
    """
    Real workbook provider backed by openpyxl.

    Access modes:
    - open_for_read:     fast streaming reads for cell/range/formula scans
    - open_for_metadata: normal workbook open for sheet metadata like merged cells
    - open_for_write:    writable workbook open
    """

    def _validate_path(self, workbook_path: str) -> Path:
        path = Path(workbook_path)

        if not workbook_path:
            raise ValueError("Missing required context: workbook_path")

        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {workbook_path}")

        if not path.is_file():
            raise FileNotFoundError(f"Workbook path is not a file: {workbook_path}")

        return path

    def open_for_read(self, workbook_path: str, *, data_only: bool = False):
        """
        Fast read path for large scans.
        Uses read_only=True for better performance.
        """
        path = self._validate_path(workbook_path)
        try:
            return load_workbook(
                filename=str(path),
                read_only=True,
                data_only=data_only,
                keep_links=False,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to open workbook for read: {workbook_path}") from exc

    def open_for_metadata(self, workbook_path: str):
        """
        Metadata path for worksheet features unavailable in read-only mode,
        such as merged_cells.
        """
        path = self._validate_path(workbook_path)
        try:
            return load_workbook(
                filename=str(path),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to open workbook for metadata: {workbook_path}") from exc

    def open_for_write(self, workbook_path: str):
        """
        Writable workbook open.
        """
        path = self._validate_path(workbook_path)
        try:
            return load_workbook(
                filename=str(path),
                read_only=False,
                data_only=False,
                keep_links=False,
            )
        except Exception as exc:
            raise RuntimeError(f"Failed to open workbook for write: {workbook_path}") from exc

    def save(self, workbook, workbook_path: str) -> None:
        path = self._validate_path(workbook_path)

        try:
            workbook.save(str(path))
        except Exception as exc:
            raise RuntimeError(f"Failed to save workbook: {workbook_path}") from exc

    def close_quietly(self, workbook) -> None:
        try:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()
        except Exception:
            pass


# -----------------------------
# Server Initialization
# -----------------------------
workbook_provider = WorkbookProvider()
excel_server = ExcelMCPServer(workbook_provider=workbook_provider)

app = FastAPI(
    title="excel-mcp",
    version="0.3.2",
)


# -----------------------------
# Request Models
# -----------------------------
class CallRequest(BaseModel):
    tool_name: str
    args: Dict[str, Any] = Field(default_factory=dict)
    ctx: Dict[str, Any] = Field(default_factory=dict)


class CallToolRequest(BaseModel):
    tool: str
    args: Dict[str, Any] = Field(default_factory=dict)
    ctx: Dict[str, Any] = Field(default_factory=dict)


# -----------------------------
# Utility
# -----------------------------
def _normalize_ctx(args: Dict[str, Any], ctx: Dict[str, Any]) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Backward compatibility:
    If an old client sends workbook_path inside args, promote it into ctx.
    """
    args_out = dict(args or {})
    ctx_out = dict(ctx or {})

    if not ctx_out.get("workbook_path") and args_out.get("workbook_path"):
        ctx_out["workbook_path"] = args_out["workbook_path"]

    return args_out, ctx_out


def safe_dispatch(tool_name: str, args: Dict[str, Any], ctx: Dict[str, Any]):
    """
    Centralized dispatch with strong error handling.
    Converts internal exceptions to proper HTTP errors.
    """
    args, ctx = _normalize_ctx(args, ctx)

    try:
        return excel_server.dispatch(tool_name, args, ctx)

    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc

    except KeyError as exc:
        missing = str(exc.args[0]) if exc.args else "unknown"
        if missing == "workbook_path":
            raise HTTPException(
                status_code=400,
                detail="Missing required context: workbook_path",
            ) from exc
        raise HTTPException(
            status_code=400,
            detail=f"Missing required argument: {missing}",
        ) from exc

    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc

    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"excel-mcp server error: {exc}",
        ) from exc


# -----------------------------
# Endpoints
# -----------------------------
@app.get("/health")
def health():
    return {"ok": True}


@app.get("/list_tools")
def list_tools():
    return {"tools": excel_server.list_tools()}


@app.post("/call")
def call(req: CallRequest):
    return safe_dispatch(req.tool_name, req.args, req.ctx)


@app.post("/call_tool")
def call_tool(req: CallToolRequest):
    return safe_dispatch(req.tool, req.args, req.ctx)